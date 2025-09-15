from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import os
import boto3
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import PyPDF2
import pdfplumber
from PIL import Image
import io
import base64
from config import Config

app = Flask(__name__)
app.secret_key = Config.SECRET_KEY

# AWS Bedrock設定
bedrock_client = boto3.client(
    service_name='bedrock-runtime',
    region_name=Config.AWS_DEFAULT_REGION,
    aws_access_key_id=Config.AWS_ACCESS_KEY_ID,
    aws_secret_access_key=Config.AWS_SECRET_ACCESS_KEY
)

class PDFCorrector:
    def __init__(self):
        self.corrections = []
    
    def extract_text_from_pdf(self, pdf_path):
        """PDFからテキストを抽出（最大3ページまで）"""
        text_content = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                max_pages = min(len(pdf.pages), Config.MAX_PDF_PAGES)
                for page_num in range(max_pages):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        text_content.append({
                            'page': page_num + 1,
                            'text': text
                        })
        except Exception as e:
            print(f"PDF読み込みエラー: {e}")
        return text_content
    
    def extract_images_from_pdf(self, pdf_path):
        """PDFから画像を抽出（最大3ページまで）"""
        images = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                max_pages = min(len(pdf.pages), Config.MAX_PDF_PAGES)
                for page_num in range(max_pages):
                    page = pdf.pages[page_num]
                    page_images = page.images
                    for img in page_images:
                        images.append({
                            'page': page_num + 1,
                            'bbox': img['bbox'],
                            'x0': img['x0'],
                            'y0': img['y0'],
                            'x1': img['x1'],
                            'y1': img['y1']
                        })
        except Exception as e:
            print(f"画像抽出エラー: {e}")
        return images
    
    def check_with_claude(self, content, content_type="text"):
        """Claude 3.5 Sonnet v2で校正チェック"""
        try:
            if content_type == "text":
                prompt = f"""
以下のテキストを校正してください。誤字脱字、文法ミス、表現の不自然さなどをチェックし、修正提案をしてください。

テキスト:
{content}

以下の形式で回答してください:
- 誤字脱字: [発見した誤字脱字とその修正案]
- 文法・表現: [文法ミスや不自然な表現とその修正案]
- その他: [その他の気になる点と修正提案]
"""
            else:
                prompt = f"""
以下の画像配置情報をチェックしてください。画像の配置ミス、重複、不適切な配置などを確認し、修正提案をしてください。

画像情報:
{content}

以下の形式で回答してください:
- 配置問題: [発見した配置の問題点]
- 修正提案: [具体的な修正案]
- その他: [その他の気になる点]
"""

            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 1000,
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ]
            })

            response = bedrock_client.invoke_model(
                modelId=Config.BEDROCK_MODEL_ID,
                contentType="application/json",
                accept="application/json",
                body=body
            )

            response_body = json.loads(response['body'].read())
            return response_body['content'][0]['text']
            
        except Exception as e:
            return f"AI校正エラー: {str(e)}"
    
    def process_pdf(self, pdf_path):
        """PDFを処理して校正結果を生成"""
        # テキスト抽出と校正
        text_content = self.extract_text_from_pdf(pdf_path)
        for item in text_content:
            correction = self.check_with_claude(item['text'], "text")
            self.corrections.append({
                'type': 'text',
                'page': item['page'],
                'content': item['text'][:100] + '...' if len(item['text']) > 100 else item['text'],
                'correction': correction
            })
        
        # 画像抽出と校正
        images = self.extract_images_from_pdf(pdf_path)
        for img in images:
            correction = self.check_with_claude(str(img), "image")
            self.corrections.append({
                'type': 'image',
                'page': img['page'],
                'content': f"画像位置: ({img['x0']}, {img['y0']}) - ({img['x1']}, {img['y1']})",
                'correction': correction
            })
        
        return self.corrections
    
    def export_to_excel(self, output_path):
        """校正結果をエクセルに出力"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "校正結果"
        
        # ヘッダー設定
        headers = ['ページ', 'タイプ', '内容', '校正結果']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # データ入力
        for row, correction in enumerate(self.corrections, 2):
            ws.cell(row=row, column=1, value=correction['page'])
            ws.cell(row=row, column=2, value=correction['type'])
            ws.cell(row=row, column=3, value=correction['content'])
            ws.cell(row=row, column=4, value=correction['correction'])
        
        # 列幅調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)

def login_required(f):
    """ログインが必要なページのデコレータ"""
    def decorated_function(*args, **kwargs):
        if 'logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user_id = request.form.get('user_id')
        password = request.form.get('password')
        
        if user_id == Config.LOGIN_ID and password == Config.LOGIN_PASSWORD:
            session['logged_in'] = True
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='IDまたはパスワードが正しくありません')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    return redirect(url_for('login'))

@app.route('/robots.txt')
def robots_txt():
    return send_file('static/robots.txt', mimetype='text/plain')

@app.route('/')
@login_required
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'ファイルが選択されていません'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'ファイルが選択されていません'}), 400
        
        if file and file.filename.lower().endswith('.pdf'):
            # 一時ファイルとして保存
            filename = f"temp_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = os.path.join('uploads', filename)
            os.makedirs('uploads', exist_ok=True)
            file.save(filepath)
            
            # PDF校正処理
            corrector = PDFCorrector()
            corrections = corrector.process_pdf(filepath)
            
            # エクセル出力
            excel_filename = f"校正結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_path = os.path.join('outputs', excel_filename)
            os.makedirs('outputs', exist_ok=True)
            corrector.export_to_excel(excel_path)
            
            # 一時ファイル削除
            os.remove(filepath)
            
            return jsonify({
                'success': True,
                'corrections': corrections,
                'excel_file': excel_filename,
                'max_pages': Config.MAX_PDF_PAGES,
                'message': f'PDF校正が完了しました（最大{Config.MAX_PDF_PAGES}ページまで処理）'
            })
        
        return jsonify({'error': 'PDFファイルをアップロードしてください'}), 400
    
    except Exception as e:
        print(f"アップロードエラー: {str(e)}")
        return jsonify({'error': f'アップロード中にエラーが発生しました: {str(e)}'}), 500

@app.route('/download/<filename>')
@login_required
def download_file(filename):
    return send_file(os.path.join('outputs', filename), as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
