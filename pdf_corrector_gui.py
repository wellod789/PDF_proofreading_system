"""
PDF校正システム - GUI版
tkinterを使用したデスクトップアプリケーション
テキスト分析と画像分析の両方に対応
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime
import webbrowser
import boto3
import json
import base64
import io
from concurrent.futures import ThreadPoolExecutor, as_completed
from pdf2image import convert_from_path
from pdf_corrector_module import PDFCorrector
from config import Config

class PDFCorrectorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("PDF校正システム - GUI版")
        self.root.geometry("805x600")    # 最適なサイズ
        self.root.minsize(800, 600)      # 最小サイズを調整
        self.root.maxsize(1400, 1000)    # 最大サイズを調整
        
        # 変数の初期化
        self.selected_file = None
        self.corrector = None
        self.corrections = []
        self.excel_file = None
        
        # 画像分析機能の変数
        self.analysis_mode = tk.StringVar(value="text")  # "text" or "image"
        self.save_images = tk.BooleanVar(value=False)  # 画像保存の選択
        self.bedrock_client = None
        self.ai_enabled = False
        
        # AWS Bedrock設定
        try:
            self.bedrock_client = boto3.client(
                service_name='bedrock-runtime',
                region_name=Config.AWS_DEFAULT_REGION,
                aws_access_key_id=Config.AWS_ACCESS_KEY_ID,
                aws_secret_access_key=Config.AWS_SECRET_ACCESS_KEY
            )
            self.ai_enabled = True
        except Exception as e:
            print(f"AWS Bedrock設定エラー: {e}")
            self.ai_enabled = False
        
        # スタイルの設定
        self.setup_styles()
        
        # UIの構築
        self.create_widgets()
        
        # ウィンドウの中央配置
        self.center_window()
        
        # ウィンドウサイズ変更のイベント設定（サイズ表示なし）
        # self.root.bind('<Configure>', self.on_window_resize)
    
    def setup_styles(self):
        """スタイルの設定"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # カスタムスタイル
        style.configure('Title.TLabel', font=('Arial', 16, 'bold'))
        style.configure('Header.TLabel', font=('Arial', 12, 'bold'))
        style.configure('Success.TLabel', foreground='green')
        style.configure('Error.TLabel', foreground='red')
        style.configure('Info.TLabel', foreground='blue')
    
    def create_widgets(self):
        """UIコンポーネントの作成"""
        # メインフレーム
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # タイトル
        title_label = ttk.Label(main_frame, text="PDF校正システム", style='Title.TLabel')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        subtitle_label = ttk.Label(main_frame, text="AI（Claude 3.5 Sonnet v2）による自動校正", 
                                 font=('Arial', 10))
        subtitle_label.grid(row=1, column=0, columnspan=3, pady=(0, 20))
        
        # ファイル選択フレーム
        file_frame = ttk.LabelFrame(main_frame, text="ファイル選択", padding="10")
        file_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.file_label = ttk.Label(file_frame, text="PDFファイルを選択してください", 
                                  font=('Arial', 10))
        self.file_label.grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        
        ttk.Button(file_frame, text="ファイルを選択", 
                  command=self.select_file).grid(row=0, column=1)
        
        # 分析モード選択フレームは非表示
        # mode_frame = ttk.LabelFrame(main_frame, text="分析モード", padding="10")
        # mode_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # ttk.Radiobutton(mode_frame, text="テキスト分析+画像分析（統合）", 
        #                variable=self.analysis_mode, value="text").grid(row=0, column=0, sticky=tk.W, padx=(0, 20))
        
        # # 画像分析と画像保存オプションは非表示
        # # if self.ai_enabled:
        # #     ttk.Radiobutton(mode_frame, text="画像分析（AI画像認識）", 
        # #                    variable=self.analysis_mode, value="image").grid(row=0, column=1, sticky=tk.W)
        # # else:
        # #     ttk.Label(mode_frame, text="画像分析（AI設定エラー）", 
        # #              foreground='red').grid(row=0, column=1, sticky=tk.W)
        
        # # # 画像保存オプション
        # # self.save_images_checkbox = ttk.Checkbutton(mode_frame, text="画像ファイルも保存する", 
        # #                                            variable=self.save_images)
        # # self.save_images_checkbox.grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=(10, 0))
        
        # 処理ボタンフレーム
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=3, column=0, columnspan=3, pady=(0, 10))
        
        self.process_button = ttk.Button(button_frame, text="校正を開始", 
                                       command=self.start_correction, state='disabled')
        self.process_button.grid(row=0, column=0, padx=(0, 10))
        
        self.download_button = ttk.Button(button_frame, text="Excelファイルをダウンロード", 
                                        command=self.download_excel, state='disabled')
        self.download_button.grid(row=0, column=1, padx=(0, 10))
        
        ttk.Button(button_frame, text="結果をクリア", 
                  command=self.clear_results).grid(row=0, column=2)
        
        # プログレスバー
        self.progress_var = tk.StringVar(value="待機中...")
        self.progress_label = ttk.Label(main_frame, textvariable=self.progress_var)
        self.progress_label.grid(row=5, column=0, columnspan=3, pady=(0, 5))
        
        self.progress_bar = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress_bar.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 結果表示フレーム
        result_frame = ttk.LabelFrame(main_frame, text="校正結果", padding="10")
        result_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        
        # 結果表示用のNotebook（タブ）
        self.notebook = ttk.Notebook(result_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 校正結果タブ
        self.result_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.result_frame, text="校正結果")
        
        # 校正結果の表示用Treeview
        columns = ('ページ', 'タイプ', '内容', '校正結果')
        self.result_tree = ttk.Treeview(self.result_frame, columns=columns, show='headings', height=8)
        
        # 列の設定
        self.result_tree.heading('ページ', text='ページ')
        self.result_tree.heading('タイプ', text='タイプ')
        self.result_tree.heading('内容', text='内容')
        self.result_tree.heading('校正結果', text='校正結果')
        
        self.result_tree.column('ページ', width=60)
        self.result_tree.column('タイプ', width=80)
        self.result_tree.column('内容', width=200)
        self.result_tree.column('校正結果', width=400)
        
        # スクロールバー
        scrollbar_y = ttk.Scrollbar(self.result_frame, orient=tk.VERTICAL, command=self.result_tree.yview)
        scrollbar_x = ttk.Scrollbar(self.result_frame, orient=tk.HORIZONTAL, command=self.result_tree.xview)
        self.result_tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        
        self.result_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar_y.grid(row=0, column=1, sticky=(tk.N, tk.S))
        scrollbar_x.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # 詳細表示タブ
        self.detail_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.detail_frame, text="詳細表示")
        
        self.detail_text = scrolledtext.ScrolledText(self.detail_frame, height=8, width=80)
        self.detail_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 結果ツリーの選択イベント
        self.result_tree.bind('<<TreeviewSelect>>', self.on_item_select)
        
        # グリッドの重み設定
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(7, weight=1)
        file_frame.columnconfigure(0, weight=1)
        result_frame.columnconfigure(0, weight=1)
        result_frame.rowconfigure(0, weight=1)
        self.result_frame.columnconfigure(0, weight=1)
        self.result_frame.rowconfigure(0, weight=1)
        self.detail_frame.columnconfigure(0, weight=1)
        self.detail_frame.rowconfigure(0, weight=1)
        
        # ステータスバー
        self.status_var = tk.StringVar(value="準備完了")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))
    
    def center_window(self):
        """ウィンドウを画面中央に配置"""
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    
    def select_file(self):
        """ファイル選択ダイアログ"""
        file_path = filedialog.askopenfilename(
            title="PDFファイルを選択",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        if file_path:
            self.selected_file = file_path
            filename = os.path.basename(file_path)
            self.file_label.config(text=f"選択されたファイル: {filename}")
            self.process_button.config(state='normal')
            self.status_var.set(f"ファイル選択完了: {filename}")
    
    def start_correction(self):
        """校正処理を開始"""
        if not self.selected_file:
            messagebox.showerror("エラー", "PDFファイルを選択してください。")
            return
        
        # ボタンを無効化
        self.process_button.config(state='disabled')
        self.download_button.config(state='disabled')
        
        # プログレスバーを開始
        self.progress_bar.start()
        self.progress_var.set("校正処理中...")
        self.status_var.set("校正処理を開始しました...")
        
        # 別スレッドで処理を実行
        thread = threading.Thread(target=self.run_correction)
        thread.daemon = True
        thread.start()
    
    def run_correction(self):
        """校正処理の実行（別スレッド）"""
        try:
            # 出力ディレクトリの作成
            os.makedirs('outputs', exist_ok=True)
            
            # 統合分析（テキスト分析 + 画像分析）を実行
            self.corrector = PDFCorrector()
            text_corrections = self.corrector.process_pdf(
                self.selected_file, 
                progress_callback=self.update_progress
            )
            
            # テキスト分析と画像分析を並列実行
            if self.ai_enabled:
                self.update_progress("テキスト分析と画像分析を並列実行中...")
                
                # 並列処理でテキスト分析と画像分析を同時実行
                with ThreadPoolExecutor(max_workers=2) as executor:
                    # テキスト分析と画像分析を同時に開始
                    text_future = executor.submit(self.run_text_analysis_only)
                    image_future = executor.submit(self.run_image_analysis)
                    
                    # 両方の完了を待機
                    text_corrections = text_future.result()
                    image_corrections = image_future.result()
                
                # AIでテキスト分析と画像分析の結果を統合
                self.update_progress("AIで結果を統合中...")
                self.corrections = self.integrate_analysis_results(text_corrections, image_corrections)
            else:
                # AI機能が無効の場合はテキスト分析のみ
                self.corrections = text_corrections
            
            # エクセルファイルの生成
            excel_filename = f"校正結果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_path = os.path.join('outputs', excel_filename)
            self.export_combined_analysis_to_excel(excel_path)
            self.excel_file = excel_filename
            
            # UIの更新（メインスレッドで実行）
            self.root.after(0, self.correction_completed)
            
        except Exception as e:
            error_msg = f"校正処理中にエラーが発生しました: {str(e)}"
            self.root.after(0, lambda: self.correction_error(error_msg))
    
    def update_progress(self, message):
        """プログレス更新（別スレッドから呼び出し）"""
        self.root.after(0, lambda: self.progress_var.set(message))
    
    def correction_completed(self):
        """校正完了時の処理"""
        self.progress_bar.stop()
        self.progress_var.set("校正完了！")
        self.status_var.set(f"校正完了 - {len(self.corrections)}件の結果")
        
        # 結果を表示
        self.display_results()
        
        # ボタンを有効化
        self.process_button.config(state='normal')
        self.download_button.config(state='normal')
        
        messagebox.showinfo("完了", f"校正が完了しました。\n{len(self.corrections)}件の校正結果が見つかりました。")
    
    def correction_error(self, error_msg):
        """校正エラー時の処理"""
        self.progress_bar.stop()
        self.progress_var.set("エラーが発生しました")
        self.status_var.set("エラーが発生しました")
        
        # ボタンを有効化
        self.process_button.config(state='normal')
        
        messagebox.showerror("エラー", error_msg)
    
    def display_results(self):
        """校正結果を表示"""
        # 既存の結果をクリア
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        
        # 結果を追加
        for correction in self.corrections:
            # タイプの表示名を決定
            if correction['type'] == 'text':
                type_name = 'テキスト'
            elif correction['type'] == 'image':
                type_name = '画像'
            elif correction['type'] == 'integrated':
                type_name = '校正'
            elif correction['type'] == 'info':
                type_name = '情報'
            else:
                type_name = correction['type']
            
            self.result_tree.insert('', 'end', values=(
                correction['page'],
                type_name,
                correction['content'],
                correction['correction'][:100] + '...' if len(correction['correction']) > 100 else correction['correction']
            ))
    
    def on_item_select(self, event):
        """結果アイテムが選択された時の処理"""
        selection = self.result_tree.selection()
        if selection:
            item = self.result_tree.item(selection[0])
            values = item['values']
            
            # 詳細表示を更新
            self.detail_text.delete(1.0, tk.END)
            detail_text = f"""
ページ: {values[0]}
タイプ: {values[1]}
内容: {values[2]}

校正結果:
{values[3]}
"""
            self.detail_text.insert(1.0, detail_text)
    
    def download_excel(self):
        """Excelファイルをダウンロード"""
        if not self.excel_file:
            messagebox.showwarning("警告", "ダウンロード可能なファイルがありません。")
            return
        
        try:
            excel_path = os.path.join('outputs', self.excel_file)
            if os.path.exists(excel_path):
                # ファイルを開く
                os.startfile(excel_path)
                self.status_var.set(f"Excelファイルを開きました: {self.excel_file}")
            else:
                messagebox.showerror("エラー", "Excelファイルが見つかりません。")
        except Exception as e:
            messagebox.showerror("エラー", f"ファイルを開けませんでした: {str(e)}")
    
    def clear_results(self):
        """結果をクリア"""
        # 結果をクリア
        for item in self.result_tree.get_children():
            self.result_tree.delete(item)
        
        self.detail_text.delete(1.0, tk.END)
        
        # 変数をリセット
        self.corrections = []
        self.excel_file = None
        self.corrector = None
        
        # ボタンを無効化
        self.download_button.config(state='disabled')
        
        # ステータスを更新
        self.status_var.set("結果をクリアしました")
        self.progress_var.set("待機中...")
    
    def run_image_analysis(self):
        """画像分析処理の実行（並列処理用）"""
        try:
            # PDFを画像に変換
            self.root.after(0, lambda: self.update_progress("PDFを画像に変換中..."))
            images = convert_from_path(self.selected_file, dpi=200)
            
            # 最大ページ数制限
            max_pages = min(len(images), Config.MAX_PDF_PAGES)
            images = images[:max_pages]
            
            # 画像保存は非表示機能のため削除
            # saved_files = []
            # if self.save_images.get():
            #     os.makedirs('converted_images', exist_ok=True)
            #     timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            #     base_name = os.path.splitext(os.path.basename(self.selected_file))[0]
            
            corrections = []
            for i, image in enumerate(images):
                page_num = i + 1
                self.root.after(0, lambda msg=f"ページ {page_num} をAI分析中...": self.update_progress(msg))
                
                # 画像をbase64エンコード
                img_buffer = io.BytesIO()
                image.save(img_buffer, format='PNG')
                img_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
                
                # 画像保存は非表示機能のため削除
                # if self.save_images.get():
                #     filename = f"{base_name}_page_{page_num:03d}_{timestamp}.png"
                #     filepath = os.path.join('converted_images', filename)
                #     image.save(filepath)
                #     saved_files.append(filepath)
                
                # AI分析
                analysis_result = self.analyze_image_with_claude(img_base64, page_num)
                corrections.append({
                    'type': 'image',
                    'page': page_num,
                    'content': f"ページ {page_num} の画像分析（視覚的問題検出含む）",
                    'correction': analysis_result
                })
            
            # 画像保存情報は非表示機能のため削除
            # if saved_files:
            #     corrections.append({
            #         'type': 'info',
            #         'page': 0,
            #         'content': f"保存された画像ファイル: {len(saved_files)}個",
            #         'correction': f"converted_images/ フォルダに保存されました。\nファイル一覧:\n" + "\n".join([os.path.basename(f) for f in saved_files])
            #     })
            
            return corrections
            
        except Exception as e:
            raise Exception(f"画像分析中にエラーが発生しました: {str(e)}")
    
    def run_text_analysis_only(self):
        """テキスト分析のみを実行（並列処理用）"""
        try:
            self.root.after(0, lambda: self.update_progress("テキスト分析を実行中..."))
            
            # PDFCorrectorを使用してテキスト分析のみ実行
            text_corrections = self.corrector.process_pdf(
                self.selected_file, 
                progress_callback=lambda msg: self.root.after(0, lambda: self.update_progress(f"テキスト分析: {msg}"))
            )
            
            return text_corrections
            
        except Exception as e:
            self.root.after(0, lambda: self.update_progress(f"テキスト分析エラー: {str(e)}"))
            return []
    
    def analyze_image_with_claude(self, image_base64, page_num):
        """Claude 3.5 Sonnet v2で画像を分析"""
        try:
            prompt = f"""
以下のPDFページ（ページ {page_num}）の画像を校正の観点から分析してください。

画像の内容について以下の観点で分析し、詳細な報告をしてください：

1. 文書の種類・目的
2. レイアウト・構成の問題点
3. テキスト内容の概要と校正すべき点
4. 画像・図表の配置と内容
5. 誤字脱字、文法ミス、表現の不自然さ
6. **視覚的な問題の検出**：
   - 不要な線、マーク、編集痕跡
   - 取り消し線、斜線、余分な図形
   - スキャン時の汚れ、ノイズ
   - 文字の重複、欠損
   - 色の不整合、コントラストの問題
7. レイアウトの改善提案
8. 全体的な校正提案

特に視覚的な問題（不要な線、編集痕跡など）については、具体的な位置と改善方法を詳しく説明してください。

日本語で回答してください。
"""

            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 2500,
                "temperature": Config.TEMPERATURE,
                "top_p": Config.TOP_P,
                "top_k": Config.TOP_K,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": prompt
                            },
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/png",
                                    "data": image_base64
                                }
                            }
                        ]
                    }
                ]
            })

            response = self.bedrock_client.invoke_model(
                modelId=Config.BEDROCK_MODEL_ID,
                contentType="application/json",
                accept="application/json",
                body=body
            )

            response_body = json.loads(response['body'].read())
            return response_body['content'][0]['text']
            
        except Exception as e:
            return f"AI分析エラー: {str(e)}"
    
    def export_image_analysis_to_excel(self, output_path):
        """画像分析結果をエクセルに出力"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "校正結果（画像分析）"
        
        # ヘッダー設定
        headers = ['ページ', 'タイプ', '内容', '校正結果']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # データ入力
        for row, correction in enumerate(self.corrections, 2):
            # タイプの表示名を決定
            if correction['type'] == 'text':
                type_name = 'テキスト'
            elif correction['type'] == 'image':
                type_name = '画像'
            elif correction['type'] == 'info':
                type_name = '情報'
            else:
                type_name = correction['type']
            
            ws.cell(row=row, column=1, value=correction['page'])
            ws.cell(row=row, column=2, value=type_name)
            ws.cell(row=row, column=3, value=correction['content'])
            ws.cell(row=row, column=4, value=correction['correction'])
        
        # 列幅調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
    
    def export_combined_analysis_to_excel(self, output_path):
        """統合分析結果（テキスト+画像）をエクセルに出力"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "校正結果"
        
        # ヘッダー設定
        headers = ['ページ', 'タイプ', '内容', '校正結果']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # データ入力
        for row, correction in enumerate(self.corrections, 2):
            # タイプの表示名を決定
            if correction['type'] == 'text':
                type_name = 'テキスト'
            elif correction['type'] == 'image':
                type_name = '画像'
            elif correction['type'] == 'integrated':
                type_name = '校正'
            elif correction['type'] == 'info':
                type_name = '情報'
            else:
                type_name = correction['type']
            
            ws.cell(row=row, column=1, value=correction['page'])
            ws.cell(row=row, column=2, value=type_name)
            ws.cell(row=row, column=3, value=correction['content'])
            ws.cell(row=row, column=4, value=correction['correction'])
        
        # 列幅調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
    
    def integrate_analysis_results(self, text_corrections, image_corrections):
        """AIでテキスト分析と画像分析の結果を統合"""
        try:
            # ページごとにグループ化
            page_groups = {}
            
            # テキスト分析結果をページごとにグループ化
            for correction in text_corrections:
                page = correction['page']
                if page not in page_groups:
                    page_groups[page] = {'text': [], 'image': []}
                page_groups[page]['text'].append(correction)
            
            # 画像分析結果をページごとにグループ化
            for correction in image_corrections:
                page = correction['page']
                if page not in page_groups:
                    page_groups[page] = {'text': [], 'image': []}
                page_groups[page]['image'].append(correction)
            
            # 各ページの結果をAIで統合
            integrated_results = []
            for page_num in sorted(page_groups.keys()):
                if page_num == 0:  # 情報項目はスキップ
                    continue
                    
                self.root.after(0, lambda msg=f"ページ {page_num} の結果を分析中...": self.update_progress(msg))
                
                text_results = page_groups[page_num]['text']
                image_results = page_groups[page_num]['image']
                
                # AIで統合
                integrated_result = self.integrate_page_results_with_ai(page_num, text_results, image_results)
                integrated_results.append(integrated_result)
            
            return integrated_results
            
        except Exception as e:
            # エラーの場合は元の結果をそのまま返す
            print(f"統合処理エラー: {e}")
            return text_corrections + image_corrections
    
    def integrate_page_results_with_ai(self, page_num, text_results, image_results):
        """AIでページのテキスト分析と画像分析結果を統合"""
        try:
            # テキスト分析結果をまとめる
            text_summary = ""
            for result in text_results:
                text_summary += f"テキスト分析: {result['correction']}\n\n"
            
            # 画像分析結果をまとめる
            image_summary = ""
            for result in image_results:
                image_summary += f"画像分析: {result['correction']}\n\n"
            
            prompt = f"""
以下のページ {page_num} のテキスト分析と画像分析の結果を統合し、重複を排除して最適化された校正結果を生成してください。

テキスト分析結果:
{text_summary}

画像分析結果:
{image_summary}

統合の指示:
1. 重複する指摘を統合し、1つの明確な指摘にまとめる
2. テキスト分析と画像分析の結果を相互補完的に統合する
3. 優先度の高い問題から順に整理する
4. 具体的で実行可能な修正提案を提供する
5. 視覚的な問題とテキストの問題を適切に組み合わせる

以下の形式で校正結果を提供してください:

【校正結果】
- 問題1: [具体的な問題と修正提案]
- 問題2: [具体的な問題と修正提案]
- 問題3: [具体的な問題と修正提案]
...

日本語で回答してください。
"""

            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 2000,
                "temperature": Config.TEMPERATURE,
                "top_p": Config.TOP_P,
                "top_k": Config.TOP_K,
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ]
            })

            response = self.bedrock_client.invoke_model(
                modelId=Config.BEDROCK_MODEL_ID,
                contentType="application/json",
                accept="application/json",
                body=body
            )

            response_body = json.loads(response['body'].read())
            integrated_correction = response_body['content'][0]['text']
            
            return {
                'type': 'integrated',
                'page': page_num,
                'content': f"ページ {page_num} の校正結果",
                'correction': integrated_correction
            }
            
        except Exception as e:
            # エラーの場合は元の結果をそのまま返す
            return {
                'type': 'integrated',
                'page': page_num,
                'content': f"ページ {page_num} の校正結果（エラーにより統合失敗）",
                'correction': f"統合処理中にエラーが発生しました: {str(e)}\n\nテキスト分析結果:\n{text_summary}\n\n画像分析結果:\n{image_summary}"
            }

def main():
    """メイン関数"""
    root = tk.Tk()
    app = PDFCorrectorGUI(root)
    
    # ウィンドウの終了処理
    def on_closing():
        if messagebox.askokcancel("終了", "アプリケーションを終了しますか？"):
            root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()

if __name__ == "__main__":
    main()
