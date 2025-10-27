"""
PDF校正モジュール
既存のapp.pyからPDFCorrectorクラスを分離して再利用可能にしたもの
"""
import os
import boto3
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
import pdfplumber
from pdf2image import convert_from_path
from PIL import Image
import io
import base64
from config import Config

class PDFCorrector:
    def __init__(self):
        self.corrections = []
        # AWS Bedrock設定
        self.bedrock_client = boto3.client(
            service_name='bedrock-runtime',
            region_name=Config.AWS_DEFAULT_REGION,
            aws_access_key_id=Config.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=Config.AWS_SECRET_ACCESS_KEY
        )
    
    def extract_text_from_pdf(self, pdf_path):
        """PDFからテキストを抽出（最大3ページまで）"""
        text_content = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                max_pages = min(len(pdf.pages), Config.MAX_PDF_PAGES)
                for page_num in range(max_pages):
                    page = pdf.pages[page_num]
                    text = page.extract_text()
                    if text:
                        text_content.append({
                            'page': page_num + 1,
                            'text': text
                        })
        except Exception as e:
            print(f"PDF読み込みエラー: {e}")
        return text_content
    
    def extract_images_from_pdf(self, pdf_path):
        """PDFから画像を抽出（最大3ページまで）"""
        images = []
        try:
            with pdfplumber.open(pdf_path) as pdf:
                max_pages = min(len(pdf.pages), Config.MAX_PDF_PAGES)
                for page_num in range(max_pages):
                    page = pdf.pages[page_num]
                    page_images = page.images
                    for img in page_images:
                        images.append({
                            'page': page_num + 1,
                            'bbox': img['bbox'],
                            'x0': img['x0'],
                            'y0': img['y0'],
                            'x1': img['x1'],
                            'y1': img['y1']
                        })
        except Exception as e:
            print(f"画像抽出エラー: {e}")
        return images
    
    def check_with_claude(self, content, content_type="text"):
        """Claude 3.5 Sonnet v2で校正チェック"""
        try:
            if content_type == "text":
                prompt = f"""
以下のテキストを校正してください。誤字脱字、文法ミス、表現の不自然さなどをチェックし、修正提案をしてください。

テキスト:
{content}

以下の形式で回答してください:
- 誤字脱字: [発見した誤字脱字とその修正案]
- 文法・表現: [文法ミスや不自然な表現とその修正案]
- その他: [その他の気になる点と修正提案]
"""
            else:
                prompt = f"""
以下の画像配置情報をチェックしてください。画像の配置ミス、重複、不適切な配置などを確認し、修正提案をしてください。

画像情報:
{content}

以下の形式で回答してください:
- 配置問題: [発見した配置の問題点]
- 修正提案: [具体的な修正案]
- その他: [その他の気になる点]
"""

            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 1000,
                "temperature": Config.TEMPERATURE,
                "top_p": Config.TOP_P,
                "top_k": Config.TOP_K,
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ]
            })

            response = self.bedrock_client.invoke_model(
                modelId=Config.BEDROCK_MODEL_ID,
                contentType="application/json",
                accept="application/json",
                body=body
            )

            response_body = json.loads(response['body'].read())
            return response_body['content'][0]['text']
            
        except Exception as e:
            return f"AI校正エラー: {str(e)}"
    
    def process_pdf(self, pdf_path, progress_callback=None):
        """PDFを処理して校正結果を生成"""
        self.corrections = []
        
        # テキスト抽出と校正
        if progress_callback:
            progress_callback("テキストを抽出中...")
        
        text_content = self.extract_text_from_pdf(pdf_path)
        for i, item in enumerate(text_content):
            if progress_callback:
                progress_callback(f"テキスト校正中... ({i+1}/{len(text_content)})")
            
            correction = self.check_with_claude(item['text'], "text")
            self.corrections.append({
                'type': 'text',
                'page': item['page'],
                'content': item['text'][:100] + '...' if len(item['text']) > 100 else item['text'],
                'correction': correction
            })
        
        # 画像抽出と校正
        if progress_callback:
            progress_callback("画像情報を抽出中...")
        
        images = self.extract_images_from_pdf(pdf_path)
        for i, img in enumerate(images):
            if progress_callback:
                progress_callback(f"画像校正中... ({i+1}/{len(images)})")
            
            correction = self.check_with_claude(str(img), "image")
            self.corrections.append({
                'type': 'image',
                'page': img['page'],
                'content': f"画像位置: ({img['x0']}, {img['y0']}) - ({img['x1']}, {img['y1']})",
                'correction': correction
            })
        
        if progress_callback:
            progress_callback("校正完了！")
        
        return self.corrections
    
    def export_to_excel(self, output_path):
        """校正結果をエクセルに出力"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "校正結果"
        
        # ヘッダー設定
        headers = ['ページ', 'タイプ', '内容', '校正結果']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # データ入力
        for row, correction in enumerate(self.corrections, 2):
            # タイプの表示名を決定
            if correction['type'] == 'text':
                type_name = 'テキスト'
            elif correction['type'] == 'image':
                type_name = '画像'
            elif correction['type'] == 'integrated':
                type_name = '校正'
            elif correction['type'] == 'info':
                type_name = '情報'
            else:
                type_name = correction['type']
            
            ws.cell(row=row, column=1, value=correction['page'])
            ws.cell(row=row, column=2, value=type_name)
            ws.cell(row=row, column=3, value=correction['content'])
            ws.cell(row=row, column=4, value=correction['correction'])
        
        # 列幅調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(output_path)
    
    def run_image_analysis(self, pdf_path):
        """画像分析処理の実行"""
        try:
            # PDFを画像に変換
            images = convert_from_path(pdf_path, dpi=200)
            
            # 最大ページ数制限
            max_pages = min(len(images), Config.MAX_PDF_PAGES)
            images = images[:max_pages]
            
            corrections = []
            for i, image in enumerate(images):
                page_num = i + 1
                
                # 画像をbase64エンコード
                img_buffer = io.BytesIO()
                image.save(img_buffer, format='PNG')
                img_base64 = base64.b64encode(img_buffer.getvalue()).decode('utf-8')
                
                # AI分析
                analysis_result = self.analyze_image_with_claude(img_base64, page_num)
                corrections.append({
                    'type': 'image',
                    'page': page_num,
                    'content': f"ページ {page_num} の画像分析",
                    'correction': analysis_result
                })
            
            return corrections
            
        except Exception as e:
            print(f"画像分析エラー: {e}")
            return []
    
    def analyze_image_with_claude(self, image_base64, page_num):
        """Claude 3.5 Sonnet v2で画像を分析"""
        try:
            prompt = f"""
以下のPDFページ（ページ {page_num}）の画像を校正の観点から分析してください。

特に以下の点を重点的にチェックしてください：
- 不要な線、マーク、編集痕跡
- レイアウトの問題
- 視覚的な不整合
- 画像の配置ミス
- フォントの不統一
- 余白の不適切な使用

以下の形式で回答してください:
- 視覚的問題: [発見した視覚的な問題]
- 修正提案: [具体的な修正案]
- その他: [その他の気になる点]

日本語で回答してください。
"""

            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 2500,
                "temperature": Config.TEMPERATURE,
                "top_p": Config.TOP_P,
                "top_k": Config.TOP_K,
                "messages": [
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": "image/png",
                                    "data": image_base64
                                }
                            },
                            {
                                "type": "text",
                                "text": prompt
                            }
                        ]
                    }
                ]
            })

            response = self.bedrock_client.invoke_model(
                modelId=Config.BEDROCK_MODEL_ID,
                contentType="application/json",
                accept="application/json",
                body=body
            )

            response_body = json.loads(response['body'].read())
            return response_body['content'][0]['text']
            
        except Exception as e:
            return f"画像分析エラー: {str(e)}"
    
    def integrate_analysis_results(self, text_corrections, image_corrections):
        """AIでテキスト分析と画像分析の結果を統合"""
        try:
            # ページごとにグループ化
            page_groups = {}
            
            # テキスト分析結果をページごとにグループ化
            for correction in text_corrections:
                page = correction['page']
                if page not in page_groups:
                    page_groups[page] = {'text': [], 'image': []}
                page_groups[page]['text'].append(correction)
            
            # 画像分析結果をページごとにグループ化
            for correction in image_corrections:
                page = correction['page']
                if page not in page_groups:
                    page_groups[page] = {'text': [], 'image': []}
                page_groups[page]['image'].append(correction)
            
            # 各ページの結果をAIで統合
            integrated_results = []
            for page_num in sorted(page_groups.keys()):
                if page_num == 0:  # 情報項目はスキップ
                    continue
                    
                text_results = page_groups[page_num]['text']
                image_results = page_groups[page_num]['image']
                
                # AIで統合
                integrated_result = self.integrate_page_results_with_ai(page_num, text_results, image_results)
                integrated_results.append(integrated_result)
            
            return integrated_results
            
        except Exception as e:
            # エラーの場合は元の結果をそのまま返す
            print(f"統合処理エラー: {e}")
            return text_corrections + image_corrections
    
    def integrate_page_results_with_ai(self, page_num, text_results, image_results):
        """AIでページのテキスト分析と画像分析結果を統合"""
        try:
            # テキスト分析結果をまとめる
            text_summary = ""
            for result in text_results:
                text_summary += f"テキスト分析: {result['correction']}\n\n"
            
            # 画像分析結果をまとめる
            image_summary = ""
            for result in image_results:
                image_summary += f"画像分析: {result['correction']}\n\n"
            
            prompt = f"""
以下のページ {page_num} のテキスト分析と画像分析の結果を統合し、重複を排除して最適化された校正結果を生成してください。

テキスト分析結果:
{text_summary}

画像分析結果:
{image_summary}

統合の指示:
1. 重複する指摘を統合し、1つの明確な指摘にまとめる
2. テキスト分析と画像分析の結果を相互補完的に統合する
3. 優先度の高い問題から順に整理する
4. 具体的で実行可能な修正提案を提供する
5. 視覚的な問題とテキストの問題を適切に組み合わせる

以下の形式で校正結果を提供してください:

【校正結果】
- 問題1: [具体的な問題と修正提案]
- 問題2: [具体的な問題と修正提案]
- 問題3: [具体的な問題と修正提案]
...

日本語で回答してください。
"""

            body = json.dumps({
                "anthropic_version": "bedrock-2023-05-31",
                "max_tokens": 2000,
                "temperature": Config.TEMPERATURE,
                "top_p": Config.TOP_P,
                "top_k": Config.TOP_K,
                "messages": [
                    {
                        "role": "user",
                        "content": prompt
                    }
                ]
            })

            response = self.bedrock_client.invoke_model(
                modelId=Config.BEDROCK_MODEL_ID,
                contentType="application/json",
                accept="application/json",
                body=body
            )

            response_body = json.loads(response['body'].read())
            integrated_correction = response_body['content'][0]['text']
            
            return {
                'type': 'integrated',
                'page': page_num,
                'content': f"ページ {page_num} の校正結果",
                'correction': integrated_correction
            }
            
        except Exception as e:
            # エラーの場合は元の結果をそのまま返す
            return {
                'type': 'integrated',
                'page': page_num,
                'content': f"ページ {page_num} の校正結果（エラーにより統合失敗）",
                'correction': f"統合処理中にエラーが発生しました: {str(e)}\n\nテキスト分析結果:\n{text_summary}\n\n画像分析結果:\n{image_summary}"
            }
